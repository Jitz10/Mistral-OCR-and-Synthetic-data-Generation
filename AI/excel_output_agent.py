import os
import json
import asyncio
from datetime import datetime
from typing import Dict, List, Any, Optional, Union
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pymongo
from pymongo import MongoClient
import chromadb
import yfinance as yf
import redis
from dotenv import load_dotenv
import autogen
from tools import (
    ReasoningTool,
    YFinanceNumberTool,
    YFinanceNewsTool,
    ArithmeticCalculationTool,
    VectorSearchRAGTool
)
from microagent import (
    ValuationRatiosAgent,
    NewsAnalysisAgent,
    ScenarioAnalysisAgent
)
from macroagent import (
    BusinessResearchAgent,
    SectorResearchAgent,
    CompanyDeepDiveAgent,
    DebtAndWorkingCapitalAgent,
    CurrentAffairsAgent,
    FuturePredictionsAgent,
    ConcallAnalysisAgent,
    RiskAnalysisAgent
)

# Load environment variables
load_dotenv()

# Constants
TODAY = datetime.now()
TODAY_STR = TODAY.strftime("%Y-%m-%d")
COMPANY_TICKER = "GANECOS.NS"
EXCEL_TEMPLATE = "templates/alpha_sage_template.xlsx"
OUTPUT_DIR = "output"

# MongoDB Configuration
MONGODB_URI = os.getenv("MONGODB_URI", "mongodb://localhost:27017/")
DB_NAME = "alphasage_chunks"
COLLECTION_NAME = "chunks"

# Redis Configuration
REDIS_URL = os.getenv("REDIS_URL", "redis://localhost:6379")
REDIS_EXPIRY = 3600  # 1 hour

# Excel Styles
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

class ExcelOutputAgent:
    """AutoGen agent for generating structured Excel outputs from AlphaSage insights."""
    
    def __init__(self):
        """Initialize the Excel Output Agent with required connections and tools."""
        self.mongo_client = MongoClient(MONGODB_URI)
        self.db = self.mongo_client[DB_NAME]
        self.collection = self.db[COLLECTION_NAME]
        
        self.redis_client = redis.from_url(REDIS_URL)
        
        self.chroma_client = chromadb.Client()
        self.collection = self.chroma_client.get_or_create_collection("alphasage")
        
        self.ticker = yf.Ticker(COMPANY_TICKER)
        
        # Initialize tools
        self.reasoning_tool = ReasoningTool()
        self.yfinance_number_tool = YFinanceNumberTool()
        self.yfinance_news_tool = YFinanceNewsTool()
        self.arithmetic_tool = ArithmeticCalculationTool()
        self.vector_search_tool = VectorSearchRAGTool()
        
        # Initialize agents
        self.macro_agents = {
            "business": BusinessResearchAgent(),
            "sector": SectorResearchAgent(),
            "deep_dive": CompanyDeepDiveAgent(),
            "debt_wc": DebtAndWorkingCapitalAgent(),
            "current_affairs": CurrentAffairsAgent(),
            "future": FuturePredictionsAgent(),
            "concall": ConcallAnalysisAgent(),
            "risk": RiskAnalysisAgent()
        }
        
        self.micro_agents = {
            "valuation": ValuationRatiosAgent(),
            "news": NewsAnalysisAgent(),
            "scenario": ScenarioAnalysisAgent()
        }
        
        # Create AutoGen agent
        self.agent = self._create_autogen_agent()
    
    def _create_autogen_agent(self) -> autogen.ConversableAgent:
        """Create and configure the AutoGen agent."""
        config_list = [{
            "model": "gpt-4",
            "api_key": os.getenv("OPENAI_API_KEY")
        }]
        
        return autogen.ConversableAgent(
            name="ExcelOutputAgent",
            llm_config={"config_list": config_list},
            system_message=f"""You are an expert financial analyst assistant specialized in creating structured Excel outputs.
Today is {TODAY_STR}. Your task is to generate comprehensive Excel reports for {COMPANY_TICKER}.
Focus on accuracy, traceability, and professional formatting."""
        )
    
    async def _get_cached_data(self, key: str) -> Optional[Dict]:
        """Retrieve cached data from Redis."""
        cached = self.redis_client.get(key)
        return json.loads(cached) if cached else None
    
    async def _cache_data(self, key: str, data: Dict, expiry: int = REDIS_EXPIRY):
        """Cache data in Redis."""
        self.redis_client.setex(key, expiry, json.dumps(data))
    
    async def _get_mongodb_chunks(self, category: Optional[str] = None) -> List[Dict]:
        """Retrieve relevant chunks from MongoDB."""
        query = {"company_name": COMPANY_TICKER}
        if category:
            query["category"] = category
        return list(self.collection.find(query))
    
    async def _get_chromadb_chunks(self, query: str, n_results: int = 5) -> List[Dict]:
        """Retrieve relevant chunks from ChromaDB."""
        results = self.collection.query(
            query_texts=[query],
            n_results=n_results
        )
        return results
    
    def _create_excel_workbook(self) -> openpyxl.Workbook:
        """Create a new Excel workbook with predefined structure."""
        wb = openpyxl.Workbook()
        
        # Create sheets
        sheets = {
            "Financials": self._setup_financials_sheet,
            "Projections": self._setup_projections_sheet,
            "Ratios": self._setup_ratios_sheet,
            "Assumptions": self._setup_assumptions_sheet,
            "Sources": self._setup_sources_sheet
        }
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create and setup each sheet
        for sheet_name, setup_func in sheets.items():
            ws = wb.create_sheet(sheet_name)
            setup_func(ws)
        
        return wb
    
    def _setup_financials_sheet(self, ws: openpyxl.worksheet.worksheet.Worksheet):
        """Setup the Financials sheet with headers and formatting."""
        headers = [
            "Metric", "FY2023", "FY2024", "FY2025E", "FY2026E",
            "YoY Growth", "CAGR", "Source", "Notes"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = BORDER
            cell.alignment = Alignment(horizontal='center')
        
        # Set column widths
        ws.column_dimensions['A'].width = 30
        for col in range(2, 10):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _setup_projections_sheet(self, ws: openpyxl.worksheet.worksheet.Worksheet):
        """Setup the Projections sheet with headers and formatting."""
        headers = [
            "Metric", "Base Case", "Bull Case", "Bear Case",
            "Confidence", "Source", "Notes"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = BORDER
            cell.alignment = Alignment(horizontal='center')
        
        # Set column widths
        ws.column_dimensions['A'].width = 30
        for col in range(2, 8):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _setup_ratios_sheet(self, ws: openpyxl.worksheet.worksheet.Worksheet):
        """Setup the Ratios sheet with headers and formatting."""
        headers = [
            "Ratio", "Current", "Industry Avg", "5Y Avg",
            "Trend", "Source", "Notes"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = BORDER
            cell.alignment = Alignment(horizontal='center')
        
        # Set column widths
        ws.column_dimensions['A'].width = 30
        for col in range(2, 7):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _setup_assumptions_sheet(self, ws: openpyxl.worksheet.worksheet.Worksheet):
        """Setup the Assumptions sheet with headers and formatting."""
        headers = [
            "Assumption", "Value", "Source", "Confidence",
            "Impact", "Notes"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = BORDER
            cell.alignment = Alignment(horizontal='center')
        
        # Set column widths
        ws.column_dimensions['A'].width = 40
        for col in range(2, 7):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    def _setup_sources_sheet(self, ws: openpyxl.worksheet.worksheet.Worksheet):
        """Setup the Sources sheet with headers and formatting."""
        headers = [
            "Source ID", "Type", "Date", "Title",
            "URL", "Relevance", "Notes"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = BORDER
            cell.alignment = Alignment(horizontal='center')
        
        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 50
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 30
    
    async def _populate_financials(self, ws: openpyxl.worksheet.worksheet.Worksheet):
        """Populate the Financials sheet with data from various sources."""
        # Get financial data from yfinance
        financials = self.ticker.financials
        balance_sheet = self.ticker.balance_sheet
        cash_flow = self.ticker.cashflow
        
        # Get projections from FuturePredictionsAgent
        future_agent = self.macro_agents["future"]
        projections = await future_agent.predict_future()
        
        # Get historical data from MongoDB
        historical_chunks = await self._get_mongodb_chunks("Financial Statements")
        
        # Process and populate data
        row = 2
        metrics = [
            ("Revenue", "Total Revenue"),
            ("EBITDA", "EBITDA"),
            ("Net Income", "Net Income"),
            ("EPS", "Basic EPS"),
            ("Operating Cash Flow", "Operating Cash Flow"),
            ("Free Cash Flow", "Free Cash Flow")
        ]
        
        for metric, yfinance_key in metrics:
            # Historical data
            for col, year in enumerate(["FY2023", "FY2024"], 2):
                value = self._extract_metric_value(historical_chunks, metric, year)
                if value:
                    ws.cell(row=row, column=col, value=value)
            
            # Projections
            for col, year in enumerate(["FY2025E", "FY2026E"], 4):
                value = self._extract_projection_value(projections, metric, year)
                if value:
                    ws.cell(row=row, column=col, value=value)
            
            # Growth calculations
            self._calculate_growth_metrics(ws, row)
            
            # Add source and notes
            self._add_source_info(ws, row, metric)
            
            row += 1
    
    def _extract_metric_value(self, chunks: List[Dict], metric: str, year: str) -> Optional[float]:
        """Extract metric value from MongoDB chunks."""
        for chunk in chunks:
            if metric in chunk["content"] and year in chunk["content"]:
                # Implement extraction logic
                pass
        return None
    
    def _extract_projection_value(self, projections: Dict, metric: str, year: str) -> Optional[float]:
        """Extract projection value from FuturePredictionsAgent output."""
        for projection in projections.get("projections", []):
            if metric in projection and year in projection:
                return projection[metric]
        return None
    
    def _calculate_growth_metrics(self, ws: openpyxl.worksheet.worksheet.Worksheet, row: int):
        """Calculate YoY growth and CAGR for a metric."""
        # YoY Growth
        current = ws.cell(row=row, column=3).value
        previous = ws.cell(row=row, column=2).value
        if current and previous and previous != 0:
            yoy_growth = (current - previous) / previous
            ws.cell(row=row, column=6, value=yoy_growth)
        
        # CAGR
        start = ws.cell(row=row, column=2).value
        end = ws.cell(row=row, column=5).value
        if start and end and start != 0:
            years = 3
            cagr = (end / start) ** (1/years) - 1
            ws.cell(row=row, column=7, value=cagr)
    
    def _add_source_info(self, ws: openpyxl.worksheet.worksheet.Worksheet, row: int, metric: str):
        """Add source information and notes to a metric."""
        # Add source
        ws.cell(row=row, column=8, value="Multiple Sources")
        
        # Add notes
        notes = f"Data compiled from financial statements, projections, and market data"
        ws.cell(row=row, column=9, value=notes)
    
    async def generate_excel_report(self, output_path: Optional[str] = None) -> str:
        """Generate a comprehensive Excel report for the company."""
        if not output_path:
            output_path = os.path.join(OUTPUT_DIR, f"{COMPANY_TICKER}_{TODAY_STR}.xlsx")
        
        # Create workbook
        wb = self._create_excel_workbook()
        
        # Populate sheets
        await self._populate_financials(wb["Financials"])
        # Add other sheet population methods here
        
        # Save workbook
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
        
        return output_path

async def main():
    """Main function to run the Excel Output Agent."""
    agent = ExcelOutputAgent()
    output_path = await agent.generate_excel_report()
    print(f"Excel report generated: {output_path}")

if __name__ == "__main__":
    asyncio.run(main()) 